import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO
from openpyxl.styles import Font, Alignment

def generate_excel(title, data):
    """
    Generates an Excel file with the given title and data.

    Args:
        title (list): A list of column headers for the Excel sheet.
        data (list of lists): A list of rows, where each row is a list of cell values.

    Returns:
        BytesIO: A buffer containing the generated Excel file.

    Example:
        title = ["ID", "Name", "Status"]
        data = [
            [1, "Alice", "Active"],
            [2, "Bob", "Inactive"]
        ]
        buffer = generate_excel(title, data)
    """
    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Tickets Report"
    
    # Style settings for headers
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers to the first row
    for col_num, column_title in enumerate(title, start=1):
        cell = sheet.cell(row=1, column=col_num, value=column_title)
        cell.font = bold_font
        cell.alignment = center_alignment
    
    # Populate the sheet with data
    for row_num, row_data in enumerate(data, start=2):
        for col_num, cell_value in enumerate(row_data, start=1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}{row_num}"] = cell_value
    
    # Save the workbook to a buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    return buffer
